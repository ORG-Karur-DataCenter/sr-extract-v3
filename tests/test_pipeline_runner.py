"""End-to-end test: runner takes a JobContext and produces an Excel file.

Uses the fake Gemini client auto-patched in conftest.
"""
import asyncio
from pathlib import Path

import pytest
import openpyxl

from api.job_context import JobContext
from api.pipeline_runner import run_pipeline_for_job


@pytest.fixture
def fixture_pdf(tmp_path):
    src = Path("tests/fixtures/tiny.pdf")
    if not src.exists():
        pytest.skip("tiny.pdf fixture missing")
    dest = tmp_path / "tiny.pdf"
    dest.write_bytes(src.read_bytes())
    return dest


@pytest.fixture
def fixture_template(tmp_path):
    import openpyxl as ox
    wb = ox.Workbook()
    ws = wb.active
    ws.append(["author", "year", "title", "population", "sample_size"])
    p = tmp_path / "template.xlsx"
    wb.save(p)
    return p


@pytest.mark.asyncio
async def test_run_pipeline_produces_xlsx(tmp_path, fixture_pdf, fixture_template):
    sandbox = tmp_path / "job-sandbox"
    (sandbox / "pdfs").mkdir(parents=True)
    (sandbox / "pdfs" / fixture_pdf.name).write_bytes(fixture_pdf.read_bytes())
    (sandbox / "templates").mkdir()
    (sandbox / "templates" / "template.xlsx").write_bytes(fixture_template.read_bytes())

    ctx = JobContext(
        job_id="t1", sandbox=sandbox, api_keys=["TEST_KEY_DO_NOT_USE"],
        model="gemini-2.0-flash", output_format="xlsx",
    )
    await run_pipeline_for_job(ctx)
    assert ctx.status == "done", f"Expected done, got {ctx.status}: {ctx.error_message}"
    assert ctx.result_path is not None
    assert Path(ctx.result_path).exists()

    wb = openpyxl.load_workbook(ctx.result_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "author" in headers
