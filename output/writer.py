"""Incremental Excel/CSV writer.

Writes one row per completed study immediately on aggregation. Never
batches at the end — this makes the pipeline crash-safe. If the process
dies mid-run, all completed studies are already on disk.

Uses openpyxl in write-only mode for the first write, then append-on-
reopen for subsequent writes. CSV mirror is append-only.
"""
from __future__ import annotations
import csv
import json
import logging
import threading
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl import Workbook, load_workbook

from config.settings import OUTPUT_DIR, OUTPUT_FORMAT

log = logging.getLogger(__name__)

# Serialize writes across workers — openpyxl is not thread-safe
_WRITE_LOCK = threading.Lock()


def _default_headers(record: dict, template_fields: list[str]) -> list[str]:
    """Template fields first, then metadata keys (underscore-prefixed) last."""
    meta = sorted(k for k in record.keys() if k.startswith("_"))
    return list(template_fields) + meta


def _serialize(value: Any) -> str:
    """Flatten dicts/lists to JSON string so Excel cells don't break."""
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


class IncrementalWriter:
    """Append-on-completion writer. One row per study."""

    def __init__(self, output_dir: Path = OUTPUT_DIR,
                 basename: str = "extracted",
                 fmt: str = OUTPUT_FORMAT):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.basename = basename
        self.fmt = fmt.lower()
        self.xlsx_path = self.output_dir / f"{basename}.xlsx"
        self.csv_path = self.output_dir / f"{basename}.csv"
        self._headers: list[str] | None = None

    def _init_files(self, headers: list[str]):
        """Create files with headers if they don't exist."""
        self._headers = headers
        if self.fmt in ("xlsx", "both") and not self.xlsx_path.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "extractions"
            ws.append(headers)
            # Freeze header row, auto-width hint
            ws.freeze_panes = "A2"
            wb.save(self.xlsx_path)
            log.info(f"Created {self.xlsx_path}")
        if self.fmt in ("csv", "both") and not self.csv_path.exists():
            with self.csv_path.open("w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerow(headers)
            log.info(f"Created {self.csv_path}")

    def append_record(self, record: dict[str, Any],
                      template_fields: list[str]) -> None:
        """Append one study's record. Thread-safe."""
        with _WRITE_LOCK:
            if self._headers is None:
                self._init_files(_default_headers(record, template_fields))
            row = [_serialize(record.get(h)) for h in self._headers]

            if self.fmt in ("xlsx", "both"):
                # Re-open, append, save. Slower than write-only but crash-safe.
                wb = load_workbook(self.xlsx_path)
                ws = wb["extractions"]
                ws.append(row)
                wb.save(self.xlsx_path)

            if self.fmt in ("csv", "both"):
                with self.csv_path.open("a", newline="", encoding="utf-8") as f:
                    csv.writer(f).writerow(row)

            log.info(f"Wrote study {record.get('_study_id', '?')} -> {self.xlsx_path.name}")

    def written_study_ids(self) -> set[str]:
        """Return study_ids already in the output file (to avoid duplicates)."""
        if not self.xlsx_path.exists():
            return set()
        try:
            wb = load_workbook(self.xlsx_path, read_only=True)
            ws = wb["extractions"]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return set()
            headers = list(rows[0])
            if "_study_id" not in headers:
                return set()
            idx = headers.index("_study_id")
            return {r[idx] for r in rows[1:] if r[idx]}
        except Exception as e:
            log.warning(f"Could not read existing output: {e}")
            return set()
